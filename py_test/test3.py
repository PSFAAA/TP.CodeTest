import json
import xlrd
import openpyxl
import datetime


# 文件内容转字典
class FileToDict:
    def __init__(self, file_name):
        self.file_name = file_name

    # csv文件
    def csv_file(self):
        cases = []
        csv_file = open(self.file_name, 'r')
        keys = csv_file.readline().strip('\n').split(',')
        for file in csv_file.readlines():
            values = file.strip('\n').split(',')
            case = dict(zip(keys, values))
            cases.append(case)
        return cases

    # xlsx文件
    def xlsx_file(self):
        wb = openpyxl.load_workbook(self.file_name)
        print('表单列表:', wb.sheetnames)
        sheet_name = input('表单名称：')
        sh = wb[sheet_name]
        cases = []
        # 获取最大行号
        rows = sh.max_row
        #  获取最大列号
        cols = sh.max_column
        for row in range(2, rows + 1):
            case = {}
            for col in range(1, cols + 1):
                xlsx_value = sh.cell(row, col).value
                xlsx_type = type(xlsx_value)
                if xlsx_type == datetime.datetime:
                    xlsx_value = xlsx_value.strftime('%Y-%m-%d')
                elif xlsx_type == int:
                    pass
                    # xlsx_value = f'{xlsx_value:<10d}'
                elif xlsx_type == float:
                    xlsx_value = f'{xlsx_value:.4f}'
                # print(xlsx_type, '', xlsx_value)
                case[sh.cell(1, col).value] = xlsx_value
                if col == cols - 1:
                    cases.append(case)
        wb.save(self.file_name)
        wb.close()
        return cases

    # xls文件
    def xls_file(self):
        wb = xlrd.open_workbook(self.file_name)
        print('表单列表:', wb.sheet_names())
        sheet_name = input('表单名称：')
        sh = wb.sheet_by_name(sheet_name)
        cases = []
        # 获取最大行号
        rows = sh.nrows
        #  获取最大列号
        cols = sh.ncols
        for row in range(1, rows):
            case = {}
            for col in range(0, cols):
                xls_value = sh.cell(row, col).value
                xls_type = sh.cell(row, col).ctype

                if xls_type == 2 and xls_value % 1 == 0:
                    xls_value = int(xls_value)
                elif xls_type == 3:
                    xls_date = xlrd.xldate_as_tuple(xls_value, 0)
                    # xls_value = xls_date.strftime('%Y/%d/%m %H:%M:%S')
                    xls_value = f'{xls_date[0]}-{xls_date[1]:>02d}-{xls_date[2]:>02d}'
                elif xls_type == 4:
                    xls_value = True if xls_value == 1 else False
                else:
                    xls_value = f'{xls_value:.4f}'
                # print(xls_type, '', xls_value)
                case[sh.cell(0, col).value] = xls_value
                if col == cols - 1:
                    cases.append(case)
        return cases

    # 获取文件名的后缀名
    def get_suffix(self, has_dot=False):
        pos = self.file_name.rfind('.')
        if 0 < pos < len(self.file_name) - 1:
            index = pos if has_dot else pos + 1
            return self.file_name[index:]
        else:
            return ''

    # 根据文件类型调用方法并输出
    def get_json(self):
        file_suffix = self.get_suffix()
        if file_suffix == 'xlsx':
            print(json.dumps(self.xlsx_file()))
        elif file_suffix == 'csv':
            print(json.dumps(self.csv_file()))
        elif file_suffix == 'xls':
            print(json.dumps(self.xls_file()))

# API结果校验
# class ApiResultCheck:
#     def __init__(self, file_name):
#         self.file_name = file_name


def main():
    r"""
    C:\Users\sh.pan.tim\Desktop\PPP\阿里巴巴2020年股票数据.xlsx
    C:\Users\sh.pan.tim\Desktop\PPP\阿里巴巴2020年股票数据.xls
    C:\Users\sh.pan.tim\Desktop\PPP\abes.csv
    """
    # ftj = FileToJson(r'C:\Users\sh.pan.tim\Desktop\PPP\阿里巴巴2020年股票数据.xls')
    ftj = FileToDict(input('文件路径：'))
    ftj.get_json()


if __name__ == '__main__':
    main()
